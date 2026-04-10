from __future__ import annotations

import io
from datetime import date
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ── Farben ────────────────────────────────────────────────────────────────────
M3_BLUE  = "1F4E79"   # dunkles Blau für Header
M3_LIGHT = "D6E4F0"   # helles Blau für alternierende Zeilen
GOLD     = "F4B942"   # Gold für Summen-/Überstunden-Spalten
RED_FONT = "C00000"   # Rot für negative Überstunden

SUM_COLS = {"Total travel time", "Total working time", "Target hours", "Overtime"}


def _fill_analysis_sheet(
    ws,
    df: pd.DataFrame,
    meta: dict[str, Any],
    holiday_label: str,
) -> None:
    """Fills a worksheet with the 8-category Analysis data."""
    d_from: date | None = meta.get("date_from")
    d_to:   date | None = meta.get("date_to")
    soll: float         = meta.get("soll_hours", 0.0)

    title_parts = ["M3 Croatia Time & Travel Report"]
    if d_from and d_to:
        title_parts.append(f"{d_from.strftime('%d.%m.%Y')} – {d_to.strftime('%d.%m.%Y')}")
    title_parts.append(f"Target: {soll:.1f} h")
    title_parts.append(f"Holidays: {holiday_label}")
    title_text = "   |   ".join(title_parts)

    n_cols = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font      = Font(bold=True, color="FFFFFF", size=12)
    tc.fill      = PatternFill("solid", fgColor=M3_BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdr_fill  = PatternFill("solid", fgColor=M3_BLUE)
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=c_idx, value=col_name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
    ws.row_dimensions[2].height = 44

    alt_fill  = PatternFill("solid", fgColor=M3_LIGHT)
    sum_fill  = PatternFill("solid", fgColor=GOLD)
    thin_side = Side(style="thin", color="CCCCCC")
    thin_bdr  = Border(bottom=thin_side, right=Side(style="thin", color="E0E0E0"))
    num_fmt   = '#,##0.00" h"'

    for r_idx, row_data in enumerate(df.itertuples(index=False), start=3):
        use_alt = (r_idx % 2 == 0)
        for c_idx, (col_name, val) in enumerate(zip(df.columns, row_data), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_bdr
            if col_name == "Employee":
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(vertical="center")
                if use_alt: cell.fill = alt_fill
            elif col_name in SUM_COLS:
                cell.fill = sum_fill
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
                is_neg = isinstance(val, (int, float)) and val < 0
                cell.font = Font(bold=True, size=10,
                    color=(RED_FONT if (col_name == "Overtime" and is_neg) else "000000"))
            else:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(size=10)
                if use_alt: cell.fill = alt_fill

    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 24 if col_name == "Employee" else 15
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"


def _fill_breakdown_standalone_sheet(
    ws,
    df_analyse: pd.DataFrame,
    meta: dict[str, Any],
    holiday_label: str,
) -> None:
    """
    Fills a Breakdown sheet with standalone computed values.
    df_analyse already contains all calculated columns; manual columns
    (Days of paid leave, Days of sick leave, Notes) are appended as empty.
    """
    GREEN       = "375623"
    GREEN_LIGHT = "E2EFDA"
    ORANGE      = "FCE4D6"

    d_from: date | None = meta.get("date_from")
    d_to:   date | None = meta.get("date_to")

    # Title row
    all_headers = list(df_analyse.columns) + ["Days of paid leave", "Days of sick leave", "Notes"]
    n_cols = len(all_headers)

    period_str = (
        f"{d_from.strftime('%d.%m.%Y')} – {d_to.strftime('%d.%m.%Y')}"
        if d_from and d_to else "–"
    )
    title_text = f"M3 Croatia – Breakdown   |   {period_str}   |   Holidays: {holiday_label}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font      = Font(bold=True, color="FFFFFF", size=12)
    tc.fill      = PatternFill("solid", fgColor=GREEN)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdr_fill  = PatternFill("solid", fgColor=GREEN)
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c_idx, h in enumerate(all_headers, start=1):
        cell = ws.cell(row=2, column=c_idx, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
    ws.row_dimensions[2].height = 44

    alt_fill  = PatternFill("solid", fgColor=GREEN_LIGHT)
    ora_fill  = PatternFill("solid", fgColor=ORANGE)
    thin_side = Side(style="thin", color="CCCCCC")
    thin_bdr  = Border(bottom=thin_side, right=Side(style="thin", color="E0E0E0"))
    num_fmt   = '#,##0.00" h"'
    eur_fmt   = '#,##0.00" €"'
    eur4_fmt  = '#,##0.0000" €"'

    OVERTIME_COLS = {"Overtime total", "Overtime travel (110%)", "Overtime work (110%)"}
    MANUAL_COLS   = {"Days of paid leave", "Days of sick leave", "Notes"}
    PAY_COLS      = {"Gross salary (€)", "Hourly rate (€)"}

    for r_idx, (_, row) in enumerate(df_analyse.iterrows(), start=3):
        use_alt = (r_idx % 2 == 0)
        for c_idx, h in enumerate(all_headers, start=1):
            val = row[h] if h in df_analyse.columns else None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_bdr

            if h == "Employee":
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(vertical="center")
                if use_alt: cell.fill = alt_fill
            elif h in MANUAL_COLS:
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if use_alt: cell.fill = alt_fill
            elif h in OVERTIME_COLS:
                cell.fill          = ora_fill
                cell.number_format = num_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                is_neg = isinstance(val, (int, float)) and val < 0
                cell.font = Font(bold=True, size=10,
                                 color=(RED_FONT if is_neg else "000000"))
            elif h == "Hourly rate (€)":
                cell.number_format = eur4_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                cell.font          = Font(size=10)
                if use_alt: cell.fill = alt_fill
            elif h in PAY_COLS:
                cell.number_format = eur_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                cell.font          = Font(size=10)
                if use_alt: cell.fill = alt_fill
            else:
                cell.number_format = num_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                cell.font          = Font(size=10)
                if use_alt: cell.fill = alt_fill

    # Column widths
    ws.column_dimensions["A"].width = 24
    for c_idx, h in enumerate(all_headers[1:], start=2):
        if h == "Notes":
            ws.column_dimensions[get_column_letter(c_idx)].width = 30
        elif h in ("Days of paid leave", "Days of sick leave"):
            ws.column_dimensions[get_column_letter(c_idx)].width = 18
        else:
            ws.column_dimensions[get_column_letter(c_idx)].width = 20

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"


def build_breakdown_standalone_bytes(
    df_analyse: pd.DataFrame,
    meta: dict[str, Any],
    holiday_label: str,
) -> bytes:
    """Standalone Breakdown sheet (computed values, no cross-sheet formulas) — for Steuerbüro."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Breakdown"
    _fill_breakdown_standalone_sheet(ws, df_analyse, meta, holiday_label)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_excel_bytes(
    df: pd.DataFrame,
    meta: dict[str, Any],
    holiday_label: str,
) -> bytes:
    """Full report: Analysis sheet + Breakdown sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"
    soll: float = meta.get("soll_hours", 0.0)

    _fill_analysis_sheet(ws, df, meta, holiday_label)

    # ── Breakdown-Sheet ───────────────────────────────────────────────────────
    _build_analyse_sheet(wb, df, soll)

    # Analysis-Sheet als aktives Blatt beim Öffnen setzen
    wb.active = wb["Breakdown"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_analyse_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, soll: float = 0.0) -> None:
    """
    Zweites Sheet 'Analyse' mit aggregierten Werten:
    Gesamtstunden, Mehrarbeit (aufgeteilt in Reise/Arbeit),
    und kombinierte Nacht-/Feiertags-Stunden.
    """
    GREEN      = "375623"   # dunkelgrün Header
    GREEN_LIGHT = "E2EFDA"  # hellgrün alternierend
    ORANGE     = "FCE4D6"   # orange für Mehrarbeit-Spalten

    ws = wb.create_sheet("Breakdown")

    # Englische Spaltennamen (müssen mit app.py übereinstimmen)
    COL_REISE_NIGHT_SH  = "Travel: nights & Sun/holidays"
    COL_REISE_NIGHT     = "Travel: nights"
    COL_REISE_SH        = "Travel: Sun/holidays"
    COL_ARBEIT_NIGHT_SH = "Work: nights & Sun/holidays"
    COL_ARBEIT_NIGHT    = "Work: nights"
    COL_ARBEIT_SH       = "Work: Sun/holidays"
    COL_SUM_REISE       = "Total travel time"
    COL_SUM_ARBEIT      = "Total working time"

    headers = [
        "Employee",
        "Travel + Working time",
        "Target hours",
        "Overtime total",
        "Overtime travel (110%)",
        "Overtime work (110%)",
        "Nights & Sun/holidays (75%)",
        "Nights (25%)",
        "Sun/holidays (50%)",
        "Days of paid leave",
        "Days of sick leave",
        "Notes",
    ]

    MANUAL_COLS = {"Days of paid leave", "Days of sick leave", "Notes"}

    # Styling
    hdr_fill  = PatternFill("solid", fgColor=GREEN)
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill  = PatternFill("solid", fgColor=GREEN_LIGHT)
    ora_fill  = PatternFill("solid", fgColor=ORANGE)
    thin_side = Side(style="thin", color="CCCCCC")
    thin_bdr  = Border(bottom=thin_side, right=Side(style="thin", color="E0E0E0"))
    num_fmt   = '#,##0.00" h"'

    # Kopfzeile
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
    ws.row_dimensions[1].height = 44

    # Auswertungs-Sheet: Titel in Zeile 1, Header in Zeile 2, Daten ab Zeile 3.
    # Analyse-Sheet:     Header in Zeile 1, Daten ab Zeile 2.
    # → Analyse-Zeile r  entspricht Auswertungs-Zeile r+1
    #
    # Auswertungs-Spalten (nach build_excel_bytes):
    #   A=Mitarbeiter, B=RT_NIGHT_SH, C=RT_NIGHT, D=RT_SH, E=RT_OTHER,
    #   F=Summe Reisezeit, G=WK_NIGHT_SH, H=WK_NIGHT, I=WK_SH, J=WK_OTHER,
    #   K=Summe Arbeitszeit, L=Soll-Stunden, M=Überstunden

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        use_alt   = (r_idx % 2 == 0)
        aw_row    = r_idx + 1   # entsprechende Zeile im Auswertungs-Sheet

        # Formeln – Excel speichert intern englische Namen; Semikolon für IF→WENN
        f_gesamt      = f"=Analysis!F{aw_row}+Analysis!K{aw_row}"
        f_mehrarbeit  = f"=B{r_idx}-C{r_idx}"
        f_meh_reise   = f"=IF(D{r_idx}<=Analysis!F{aw_row},D{r_idx},Analysis!F{aw_row})"
        f_meh_arbeit  = f"=D{r_idx}-E{r_idx}"
        f_night_sh    = f"=Analysis!G{aw_row}+Analysis!B{aw_row}"
        f_night       = f"=Analysis!H{aw_row}+Analysis!C{aw_row}"
        f_sun_hol     = f"=Analysis!I{aw_row}+Analysis!D{aw_row}"

        # Per-employee target hours (from "Effektive Arbeitsstunden" if present, else global soll)
        emp_soll = row["Target hours"] if "Target hours" in df.columns else soll

        values = [
            row["Employee"],  # A – Wert, kein Formel nötig
            f_gesamt,         # B
            emp_soll,         # C – per-employee target hours
            f_mehrarbeit,     # D
            f_meh_reise,      # E
            f_meh_arbeit,     # F
            f_night_sh,       # G
            f_night,          # H
            f_sun_hol,        # I
            None,             # J – Days of paid leave (manuell)
            None,             # K – Days of sick leave (manuell)
            None,             # L – Notes (manuell)
        ]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_bdr

            if c_idx == 1:  # Employee
                cell.font      = Font(bold=True, size=10)
                cell.alignment = Alignment(vertical="center")
                if use_alt:
                    cell.fill = alt_fill

            elif headers[c_idx - 1] in MANUAL_COLS:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font      = Font(size=10)
                if use_alt:
                    cell.fill = alt_fill

            elif headers[c_idx - 1] in ("Overtime total", "Overtime travel (110%)", "Overtime work (110%)"):
                cell.fill          = ora_fill
                cell.number_format = num_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                is_neg = isinstance(val, (int, float)) and val < 0
                cell.font = Font(bold=True, size=10, color=(RED_FONT if is_neg else "000000"))

            else:
                cell.number_format = num_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                cell.font          = Font(size=10)
                if use_alt:
                    cell.fill = alt_fill

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 24
    for c_idx, h in enumerate(headers[1:], start=2):
        if h == "Notes":
            ws.column_dimensions[get_column_letter(c_idx)].width = 30
        elif h in ("Days of paid leave", "Days of sick leave"):
            ws.column_dimensions[get_column_letter(c_idx)].width = 18
        else:
            ws.column_dimensions[get_column_letter(c_idx)].width = 20

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
